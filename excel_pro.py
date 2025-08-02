
import os
import pandas as pd
import re
import fitz  # PyMuPDF
import numpy as np
import uuid
from typing import List, Dict, Tuple, Any, Optional
from sentence_transformers import SentenceTransformer, CrossEncoder
import chromadb
from chromadb import PersistentClient
import boto3
import json
import camelot
import tabula
import time
from typing import List, Dict, Any
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.cell.cell import MergedCell
# Add this at the top of your script to suppress the specific warning
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="cryptography.hazmat.primitives.ciphers.algorithms")

# --- Configuration ---
EMBEDDING_MODEL_NAME = 'BAAI/bge-small-en-v1.5' # Good general-purpose embedding model
RERANKER_MODEL_NAME = 'cross-encoder/ms-marco-MiniLM-L-6-v2' # Effective reranking model
CHUNK_SIZE_TARGET = 500 # Target number of characters per chunk (approximate)
CHUNK_OVERLAP = 100    # Number of characters to overlap between chunks

# AWS Bedrock Configuration
AWS_REGION = 'ap-south-1'  # Change to your preferred region
BEDROCK_MODEL_ID = 'anthropic.claude-3-haiku-20240307-v1:0'  

def initialize_bedrock_client():
    """Initialize AWS Bedrock Runtime client"""
    try:
        bedrock_runtime = boto3.client(
            service_name='bedrock-runtime',
            region_name=AWS_REGION
        )
        print(f"AWS Bedrock client initialized successfully for region: {AWS_REGION}")
        return bedrock_runtime
    except Exception as e:
        print(f"Error initializing Bedrock client: {e}")
        print("Please ensure your AWS credentials are configured correctly.")
        return None

# Part 1: Text Extraction with Structure Preservation - Improved Version

def extract_structured_text_from_excel(file_path: str) -> List[Dict[str, Any]]:
    """
    Extract logical sections and tables from a structured Excel file with improved detection.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        List of structured blocks with type, section, headers, and rows
    """
    try:
        wb = load_workbook(file_path, data_only=True)
        structured_blocks = []
        
        print(f"Processing Excel file with {len(wb.worksheets)} worksheets")
        
        for sheet_index, sheet in enumerate(wb.worksheets):
            print(f"  Processing worksheet: {sheet.title}")
            
            # Track section information
            current_section = None
            current_headers = []
            current_rows = []
            collecting_table = False
            row_count = 0
            
            # Debug information about the sheet
            max_row = sheet.max_row
            max_col = sheet.max_column
            print(f"    Sheet dimensions: {max_row} rows x {max_col} columns")
            
            for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
                row_count += 1
                
                # Normalize row: strip whitespace from strings and convert None to empty string
                normalized_row = []
                for cell in row:
                    if cell is None:
                        normalized_row.append("")
                    elif isinstance(cell, str):
                        normalized_row.append(cell.strip())
                    else:
                        # Keep numbers and other data types as is
                        normalized_row.append(cell)
                
                # Count non-empty cells for detection logic
                non_empty_cells = [cell for cell in normalized_row if cell not in [None, ""]]
                non_empty_count = len(non_empty_cells)
                
                # Skip completely empty rows, but handle table finalization
                if non_empty_count == 0:
                    if collecting_table and current_headers and current_rows:
                        # An empty row might indicate the end of a table
                        structured_blocks.append({
                            "type": "table",
                            "section": current_section,
                            "headers": current_headers,
                            "rows": current_rows,
                            "sheet": sheet.title,  # Add sheet name for better traceability
                            "row_range": f"{row_index - len(current_rows) - 1}-{row_index - 1}"  # Track row position
                        })
                        print(f"    Found table with {len(current_rows)} rows in section '{current_section}'")
                        current_headers = []
                        current_rows = []
                        collecting_table = False
                    continue
                
                # Possible section header: one non-empty cell that's a string
                if non_empty_count == 1 and isinstance(non_empty_cells[0], str):
                    # If we were collecting a table, finalize it
                    if collecting_table and current_headers and current_rows:
                        structured_blocks.append({
                            "type": "table",
                            "section": current_section,
                            "headers": current_headers,
                            "rows": current_rows,
                            "sheet": sheet.title,
                            "row_range": f"{row_index - len(current_rows) - 1}-{row_index - 1}"
                        })
                        print(f"    Found table with {len(current_rows)} rows in section '{current_section}'")
                        current_headers = []
                        current_rows = []
                    
                    # Set the new section
                    current_section = non_empty_cells[0]
                    collecting_table = False
                    print(f"    Found section: '{current_section}'")
                    
                    # Add section as a separate block for better searchability
                    structured_blocks.append({
                        "type": "section",
                        "section": current_section,
                        "text": current_section,
                        "sheet": sheet.title,
                        "row": row_index
                    })
                    continue
                
                # Detect table header row - after section, with multiple non-empty cells
                if non_empty_count >= 2 and (not collecting_table or not current_headers):
                    # Check if this looks like a header row (contains string values, not just numbers)
                    has_string_values = any(isinstance(cell, str) and cell.strip() for cell in non_empty_cells)
                    
                    if has_string_values:
                        # Finalize previous table if we were collecting one
                        if collecting_table and current_headers and current_rows:
                            structured_blocks.append({
                                "type": "table",
                                "section": current_section,
                                "headers": current_headers,
                                "rows": current_rows,
                                "sheet": sheet.title,
                                "row_range": f"{row_index - len(current_rows) - 1}-{row_index - 1}"
                            })
                            print(f"    Found table with {len(current_rows)} rows in section '{current_section}'")
                            current_rows = []
                        
                        # Start new table with these headers
                        current_headers = []
                        for cell in normalized_row:
                            if isinstance(cell, str):
                                current_headers.append(cell)
                            else:
                                # Convert non-string headers to strings
                                current_headers.append(str(cell) if cell is not None else "")
                        
                        collecting_table = True
                        print(f"    Found table headers: {', '.join(current_headers)}")
                        continue
                
                # Collect data rows for tables
                if collecting_table and current_headers:
                    # Convert all values to string for consistency
                    row_values = []
                    for cell in normalized_row:
                        if cell is None:
                            row_values.append("")
                        elif isinstance(cell, (int, float, bool)):
                            row_values.append(str(cell))
                        elif isinstance(cell, str):
                            row_values.append(cell)
                        else:
                            # Handle other types like dates
                            row_values.append(str(cell))
                    
                    # Only add rows that have at least some content
                    if any(val.strip() for val in row_values if isinstance(val, str)):
                        current_rows.append(row_values)
            
            # End of sheet - finalize any remaining table
            if collecting_table and current_headers and current_rows:
                structured_blocks.append({
                    "type": "table",
                    "section": current_section,
                    "headers": current_headers,
                    "rows": current_rows,
                    "sheet": sheet.title,
                    "row_range": f"{row_count - len(current_rows)}-{row_count}"
                })
                print(f"    Found table with {len(current_rows)} rows in section '{current_section}'")
        
        print(f"Extraction complete. Found {len(structured_blocks)} structured blocks.")
        return structured_blocks
        
    except Exception as e:
        print(f"Error extracting data from Excel file: {e}")
        import traceback
        traceback.print_exc()
        return []


# Part 2: Structural Chunking
def create_structural_chunks(structured_blocks: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Create chunks based on document structure, grouping related blocks.
    Assigns section/subsection metadata based on headings encountered.
    """
    chunks = []
    current_section = {"heading": "Document Start", "level": 0}
    current_subsection = {"heading": "", "level": 0}
    current_sub_subsection = {"heading": "", "level": 0} # Track potentially deeper levels

    # Buffer to accumulate paragraph content within a section/subsection
    paragraph_buffer = []
    buffer_page = None

    def flush_paragraph_buffer():
        """Helper to create a chunk from accumulated paragraph content."""
        nonlocal paragraph_buffer # Need to modify the outer scope variable
        if not paragraph_buffer:
            return

        # For Excel data, we need to handle the different structure
        if paragraph_buffer[0].get("type") == "table":
            # For tables, construct a text representation
            table_data = paragraph_buffer[0]
            section = table_data.get("section", "")
            headers = table_data.get("headers", [])
            rows = table_data.get("rows", [])
            
            # Create a text representation of the table
            text_lines = []
            if section:
                text_lines.append(f"Section: {section}")
            
            # Add headers
            if headers:
                text_lines.append(" | ".join(headers))
                text_lines.append("-" * (sum(len(h) for h in headers) + 3 * (len(headers) - 1)))
            
            # Add rows
            for row in rows:
                text_lines.append(" | ".join(str(cell) for cell in row))
            
            combined_text = "\n".join(text_lines)
            
            chunks.append({
                "text": combined_text,
                "metadata": {
                    "section": current_section["heading"],
                    "subsection": current_subsection["heading"],
                    "sub_subsection": current_sub_subsection["heading"],
                    "type": "table_content",
                    "page": buffer_page # Associate chunk with starting page
                }
            })
        else:
            # For text content from PDF, use the original logic
            combined_text = "\n".join([b.get("text", "") for b in paragraph_buffer])
            
            # Determine the most relevant page number for the buffer
            buffer_first_page = paragraph_buffer[0].get("page") if paragraph_buffer else buffer_page
            
            chunks.append({
                "text": combined_text,
                "metadata": {
                    "section": current_section["heading"],
                    "subsection": current_subsection["heading"],
                    "sub_subsection": current_sub_subsection["heading"],
                    "type": "paragraph_content",
                    "page": buffer_first_page # Associate chunk with its starting page
                }
            })
        
        paragraph_buffer = []

    for i, block in enumerate(structured_blocks):
        block_type = block.get("type", "")
        
        # Handle different block types based on source (Excel vs PDF)
        if block_type == "table":
            # For Excel tables, add them directly to paragraph buffer
            paragraph_buffer.append(block)
            flush_paragraph_buffer()  # Flush immediately after adding a table
            buffer_page = None
        elif block_type.startswith("heading_level_"):
            # Handle headings (from PDF source)
            flush_paragraph_buffer()  # Flush content before the new heading
            
            block_text = block.get("text", "")
            block_page = block.get("page")
            level = int(block_type.split("_")[-1])
            
            # Update section/subsection based on heading level
            if level == 1:
                current_section = {"heading": block_text, "level": level}
                current_subsection = {"heading": "", "level": 0}
                current_sub_subsection = {"heading": "", "level": 0}
            elif level == 2:
                current_subsection = {"heading": block_text, "level": level}
                current_sub_subsection = {"heading": "", "level": 0}
            elif level == 3:
                current_sub_subsection = {"heading": block_text, "level": level}
            # For levels > 3, they are sub_sub_sub sections, just update the current_sub_subsection name
            elif level > 3:
                 # Only update if the previous level tracking was higher or equal, otherwise might mess up hierarchy
                 if current_sub_subsection["level"] == 0 or level > current_sub_subsection["level"]:
                    current_sub_subsection = {"heading": block_text, "level": level}
            
            # Optional: Add heading as a separate, small chunk for direct retrieval
            chunks.append({
                "text": block_text,
                "metadata": {
                    "section": current_section["heading"],
                    "subsection": current_subsection["heading"] if level > 1 else "",
                    "sub_subsection": current_sub_subsection["heading"] if level > 2 else "",
                    "type": block_type, # Keep the specific heading type
                    "page": block_page
                }
            })
            buffer_page = None # Reset buffer page after a heading
        elif block_type == "paragraph":
            # Handle paragraph content (from PDF source)
            paragraph_buffer.append(block)
            # Simple buffer flushing logic
            buffer_text_length = sum(len(b.get("text", "")) for b in paragraph_buffer)
            next_block_is_heading = (i + 1 < len(structured_blocks) and
                                    structured_blocks[i+1].get("type", "").startswith("heading_level_"))
            
            if buffer_text_length > CHUNK_SIZE_TARGET or next_block_is_heading:
                flush_paragraph_buffer()
                buffer_page = None # Reset buffer page after flushing

    # Flush any remaining content at the end of the document
    flush_paragraph_buffer()

    return chunks

# Part 4: Vector Database Storage (Uses the refined chunks)
def store_chunks_in_vector_db(chunks: List[Dict[str, Any]], db_path: str = "./chroma_db",
                            collection_name: str = "insurance_policy",
                            model_name: str = EMBEDDING_MODEL_NAME) -> SentenceTransformer:
    """
    Store chunks in a vector database with their embeddings.
    Uses the SentenceTransformer model specified.
    """
    print(f"Initializing embedding model: {model_name}")
    model = SentenceTransformer(model_name)

    print(f"Initializing ChromaDB client at: {db_path}")
    chroma_client = PersistentClient(path=db_path)

    print(f"Getting or creating collection: {collection_name}")
    collection = chroma_client.get_or_create_collection(name=collection_name)

    # Prepare data for storage
    texts = []
    metadatas = []
    ids = []

    for chunk in chunks:
        texts.append(chunk["text"])
        metadatas.append(sanitize_metadata(chunk["metadata"]))
        # Generate a unique ID for each chunk
        ids.append(str(uuid.uuid4()))

    if not texts:
        print("No chunks to store.")
        return model

    print(f"Generating embeddings for {len(texts)} chunks and adding to database...")
    # Generate embeddings in batches to optimize GPU usage
    batch_size = 32 # Adjust based on your GPU memory
    for i in range(0, len(texts), batch_size):
        end_idx = min(i + batch_size, len(texts))
        batch_texts = texts[i:end_idx]

        # Ensure all metadata in batch is sanitized
        batch_metadatas = []
        for meta in metadatas[i:end_idx]:
            batch_metadatas.append(meta)  # Already sanitized above

        batch_ids = ids[i:end_idx]

        # Generate embeddings
        try:
            batch_embeddings = model.encode(batch_texts, convert_to_tensor=True)
            # Convert tensor to numpy, then to list for ChromaDB
            batch_embeddings_list = batch_embeddings.cpu().numpy().tolist()
        except Exception as e:
            print(f"Error generating embeddings for batch {i//batch_size}: {e}")
            print("Skipping this batch.")
            continue # Skip adding this batch if embedding fails

        # Add to database with detailed error handling
        try:
            # Debug output for the first batch to help identify issues
            if i == 0:
                print(f"Sample metadata for first document: {batch_metadatas[0]}")

            collection.add(
                documents=batch_texts,
                metadatas=batch_metadatas,
                ids=batch_ids,
                embeddings=batch_embeddings_list
            )
            print(f"Successfully added batch {i//batch_size} ({len(batch_texts)} documents)")
        except Exception as e:
            print(f"Error adding batch {i//batch_size} to ChromaDB: {e}")
            print("Attempting to add documents individually...")

            # Try adding documents one by one to identify problematic ones
            for j in range(len(batch_texts)):
                try:
                    collection.add(
                        documents=[batch_texts[j]],
                        metadatas=[batch_metadatas[j]],
                        ids=[batch_ids[j]],
                        embeddings=[batch_embeddings_list[j]]
                    )
                except Exception as ex:
                    print(f" - Failed to add document {j} in batch {i//batch_size}: {ex}")
                    print(f" - Problematic metadata: {batch_metadatas[j]}")

    print("Finished storing chunks.")
    return model


# Part 5: Query Processing and Retrieval (Uses the defined models)
def retrieve_relevant_chunks(query: str,
                              model: SentenceTransformer,
                              db_path: str = "./chroma_db",
                              collection_name: str = "insurance_policy",
                              top_k: int = 15,
                              rerank: bool = True,
                              verbose: bool = True) -> list:
    """
    Retrieve top-k relevant chunks using embeddings and reranking.
    """
    # Connect to ChromaDB
    chroma_client = PersistentClient(path=db_path)
    try:
        collection = chroma_client.get_collection(name=collection_name)
    except:
        print(f"Error: Collection '{collection_name}' not found in '{db_path}'. Please run indexing first.")
        return [] # Return empty list if collection doesn't exist

    # Create query embedding
    query_embedding = model.encode(query).tolist()

    # Retrieve initial candidates
    # Retrieve more candidates than top_k if reranking is enabled
    initial_k = min(50, top_k * 3) if rerank else top_k
    # Ensure we don't request more results than are in the collection
    count = collection.count()
    if count == 0:
        print("Database is empty.")
        return []
    initial_k = min(initial_k, count)


    if verbose:
        print(f"Retrieving initial {initial_k} candidates...")

    results = collection.query(
        query_embeddings=[query_embedding],
        n_results=initial_k,
        include=["documents", "metadatas", "distances"] # Include distances for potential analysis
    )

    retrieved_chunks = []
    if results and results["documents"] and results["documents"][0]:
        for i in range(len(results["documents"][0])):
            retrieved_chunks.append({
                "text": results["documents"][0][i],
                "metadata": results["metadatas"][0][i],
                "distance": results["distances"][0][i] # Include distance
            })

    if verbose:
        print(f"Retrieved {len(retrieved_chunks)} initial chunks.")
        # print("\n[Debug] Top initial retrieved chunks (first 300 chars):")
        # for i, chunk in enumerate(retrieved_chunks[:min(5, len(retrieved_chunks))]):
        #     print(f"\n--- Chunk {i+1}, Distance: {chunk['distance']:.4f} ---")
        #     print(f"Section: {chunk['metadata'].get('section', 'N/A')} > {chunk['metadata'].get('subsection', 'N/A')}")
        #     print(f"Page: {chunk['metadata'].get('page', 'N/A')}")
        #     print(chunk['text'][:300] + "...\n")

    # Cross-Encoder Reranking
    if rerank and len(retrieved_chunks) > top_k:
        if verbose:
            print(f"Reranking {len(retrieved_chunks)} chunks...")
        try:
            reranker = CrossEncoder(RERANKER_MODEL_NAME)
            pairs = [(query, chunk["text"]) for chunk in retrieved_chunks]
            scores = reranker.predict(pairs, batch_size=32)

            scored_chunks = list(zip(retrieved_chunks, scores))
            scored_chunks.sort(key=lambda x: x[1], reverse=True) # Sort by reranker score

            retrieved_chunks = [chunk for chunk, _ in scored_chunks[:top_k]]

            if verbose:
                 print(f"Finished reranking, keeping top {top_k}.")
                 # print("\n[Debug] Top reranked chunks (first 300 chars):")
                 # for i, chunk in enumerate(retrieved_chunks):
                 #    print(f"\n--- Reranked Chunk {i+1} ---")
                 #    print(f"Section: {chunk['metadata'].get('section', 'N/A')} > {chunk['metadata'].get('subsection', 'N/A')}")
                 #    print(f"Page: {chunk['metadata'].get('page', 'N/A')}")
                 #    print(chunk['text'][:300] + "...\n")

        except Exception as e:
            print(f"Error during reranking: {e}. Skipping reranking.")
            # Fallback to returning the top_k from initial retrieval based on distance
            retrieved_chunks.sort(key=lambda x: x["distance"])
            retrieved_chunks = retrieved_chunks[:top_k]


    return retrieved_chunks


# Part 6: Knowledge Filtering and Context Assembly
def assemble_context(query: str,
                     chunks: List[Dict[str, Any]],
                     token_budget: int = 10000) -> str:
    """
    Assemble a context from retrieved chunks, prioritizing informative chunks
    while staying within the token budget. Includes metadata in the context.
    """
    if not chunks:
        return "No relevant information found in the policy document."
    
    def sort_key(chunk):
        type_order = {
            "heading_level_1": 1,
            "heading_level_2": 2,
            "heading_level_3": 3,
            "heading_level_4": 4,
            "paragraph_content": 5,
            "table_content": 6,
        }.get(chunk["metadata"].get("type", "paragraph_content"), 99)
        return (chunk["metadata"].get("page", 0), type_order)
    
    chunks.sort(key=sort_key)
    
    def estimate_tokens(text):
        return len(text.split()) * 1.3
    
    context_parts = []
    current_tokens = 0
    query_tokens = estimate_tokens(query)
    prompt_overhead_tokens = 500
    available_budget = token_budget - query_tokens - prompt_overhead_tokens
    
    if available_budget <= 0:
        print("Warning: Token budget is too small to include context with the query and prompt overhead.")
        available_budget = 100
    
    for chunk in chunks:
        metadata_line = f"[Page {chunk['metadata'].get('page', 'N/A')}]"
        
        section_info = []
        if chunk["metadata"].get("section"):
            section_info.append(chunk["metadata"]["section"])
        if chunk["metadata"].get("subsection"):
            section_info.append(chunk["metadata"]["subsection"])
        if chunk["metadata"].get("sub_subsection"):
            section_info.append(chunk["metadata"]["sub_subsection"])
        
        if section_info:
            metadata_line += " [" + " > ".join(section_info) + "]"
        
        text = chunk["text"]
        
        # === Improved table cleanup logic ===
        if text.count("|") > 10:  # Detect tables via pipe density
            lines = text.split("\n")
            processed_lines = []
            
            for line in lines:
                # Skip table formatting lines
                if line.strip().replace("-", "").replace("|", "") == "":
                    continue
                
                if "|" in line:
                    # Process table row
                    cells = [cell.strip() for cell in line.split("|") if cell.strip()]
                    
                    if len(cells) >= 2:
                        # Format as structured key-value pairs with clear separation
                        if ":" in cells[0]:  # If first cell already has a number/key pattern
                            # Handle numbered items like "5: Claim Preparation Cost"
                            processed_lines.append(f"{cells[0]} | " + " | ".join(cells[1:]))
                        else:
                            processed_lines.append(f"{cells[0]} | " + " | ".join(cells[1:]))
                    else:
                        processed_lines.append(" ".join(cells))
                else:
                    processed_lines.append(line)
            
            # Join with line breaks to preserve table structure visually
            joined_lines = "\n".join(processed_lines)
            formatted_chunk = f"{metadata_line}\n{joined_lines}"
        else:
            formatted_chunk = f"{metadata_line}\n{text}"
        
        chunk_tokens = estimate_tokens(formatted_chunk)
        
        if current_tokens + chunk_tokens > available_budget:
            if context_parts:
                break
            else:
                max_chars = int((available_budget - current_tokens) * 3.5)
                truncated_text = text[:max_chars]
                formatted_chunk = f"{metadata_line}\n{truncated_text}...\n[Context truncated due to length]"
                context_parts.append(formatted_chunk)
                current_tokens += estimate_tokens(formatted_chunk)
                break
        else:
            current_tokens += chunk_tokens
            context_parts.append(formatted_chunk)
    
    if not context_parts:
        return "Relevant information was retrieved, but could not be included in the context due to token budget constraints."
    
    return "\n\n".join(context_parts)

# Part 7: LLM Response Generation
# Reuse the existing generate_response function, just update the n_ctx
def generate_response_with_bedrock(query: str, context: str, bedrock_client) -> str:
    """Generate a response using AWS Bedrock with the provided context"""
    if not bedrock_client:
        return "Error: Bedrock client not initialized. Please check your AWS configuration."

    system_prompt = """You are an expert insurance underwriter with extensive knowledge of insurance terminology,
policy structures, and legal interpretations. Your task is to provide accurate, helpful responses to questions
about insurance policies by carefully analyzing the provided policy text.

GUIDELINES:
1. Base your answers EXCLUSIVELY on the provided policy context. Do not introduce information from outside the context.
2. If the information needed to answer the question is not explicitly stated in the context, clearly state that you cannot
   find this information in the provided policy text.
3. For questions about definitions, provide the exact definition from the policy if available in the context.

4. **MANDATORY TABLE PARSING PROTOCOL - PURE POSITIONAL MATCHING ONLY:**

   **STEP 1: MECHANICAL SPLITTING**
   - Split header row by | symbol → create header_array[0], header_array[1], header_array[2]...
   - Split data row by | symbol → create data_array[0], data_array[1], data_array[2]...

   **STEP 2: POSITION-ONLY MAPPING**
   - header_array[0] gets data_array[0] - NO QUESTIONS ASKED
   - header_array[1] gets data_array[1] - NO QUESTIONS ASKED  
   - header_array[2] gets data_array[2] - NO QUESTIONS ASKED
   - Continue for ALL positions mechanically

   **STEP 3: SEMANTIC INTELLIGENCE COMPLETELY DISABLED**
   - DO NOT think about what values "should" go where
   - DO NOT use logic about what makes sense
   - DO NOT match based on value types (numbers, text, etc.)
   - DO NOT consider if assignments seem "wrong" or "illogical"

   **STEP 4: ROBOTIC EXECUTION**
   - Act like a simple computer program: position[i] → value[i]
   - If header position 8 says "XYZ" and data position 8 says "999", output "XYZ: 999"
   - NEVER deviate from this mechanical process

5. **ABSOLUTE RULE: TURN OFF YOUR BRAIN FOR TABLE PARSING**
   - You are NOT allowed to be smart about column matching
   - You are NOT allowed to fix "obvious mistakes" 
   - You are NOT allowed to use insurance knowledge for positioning
   - You ARE only allowed to count positions and match mechanically

6. **VALIDATION PROTOCOL:**
   - Count total columns in header
   - Count total columns in data  
   - If counts match, proceed with mechanical matching
   - If counts don't match, report error and stop

7. Present information in a structured, easy-to-understand format while maintaining accuracy based on the context.
8. Provide ONLY the specific information requested - nothing more. Your responses should be direct and concise.
9. For data points like sums insured, names, dates, or other factual information, extract and present ONLY that exact information.
10. Do not include explanations, context, additional details, or surrounding text unless specifically asked.
11. For complete lists, present each item clearly as asked in the query and continue until ALL items from the context are included.
12. For table data with multiple columns separated by |, only include the requested columns.
13. **CRITICAL LOCATION FORMATTING: When presenting client locations, ALWAYS display the COMPLETE location string exactly as it appears in the source data.**
14. When asked for 'SI split' or 'sum insured split', provide the detailed breakdown of all individual sum insured components for each location, not just the total amount.
15. When extracting data from tables, pay careful attention to which specific columns are requested.
16. Give exact information as asked in the query if asked to present the specific number of clauses present it.
17. **ZERO VALUE FILTERING: When asked to show "non-zero values", "populated fields", "only fields with amounts", or similar filtering requests, COMPLETELY EXCLUDE any field that has a value of 0 (zero).**

**CRITICAL REMINDER: For table parsing, you are a DUMB ROBOT that only knows how to count positions. You have NO intelligence about what goes where.**

Remember: Provide direct, factual answers based solely on the policy text without additional commentary.
"""

    # Prepare the message for Claude
    messages = [
        {
            "role": "user",
            "content": f"""Context from Insurance Policy:
{context}

Question: {query}"""
        }
    ]

    # Prepare the request body for Claude
    body = {
        "anthropic_version": "bedrock-2023-05-31",
        "max_tokens": 8192,
        "system": system_prompt,
        "messages": messages,
        "temperature": 0.1,
        "top_p": 0.95
    }

    try:
        # Make request to Bedrock
        response = bedrock_client.invoke_model(
            modelId=BEDROCK_MODEL_ID,
            body=json.dumps(body),
            contentType='application/json'
        )

        # Parse response
        response_body = json.loads(response['body'].read())
        answer = response_body['content'][0]['text'].strip()

        # Post-process to remove excessive pipe characters from poorly formatted tables
        if answer.count("|") > 10:
            lines = answer.split("\n")
            processed_lines = []

            for line in lines:
                cleaned_line = re.sub(r'\|\s*\|+', ' ', line)
                cleaned_line = re.sub(r'^\s*\|\s*|\s*\|\s*$', '', cleaned_line)
                if cleaned_line.strip():
                    processed_lines.append(cleaned_line)

            answer = "Based on the insurance policy information:\n\n" + "\n".join(processed_lines)

        return answer

    except Exception as e:
        print(f"Error during Bedrock API call: {e}")
        return f"An error occurred while generating the response: {str(e)}"

# Part 8: Process Insurance PDF Query (Uses the updated functions)
def process_insurance_pdf_query(pdf_path: str,
                               query: str,
                               bedrock_client,  # Changed from model_path
                               db_path: str = "./chroma_db",
                               collection_name: str = "insurance_policy",
                               force_reindex: bool = False) -> Dict[str, Any]:
    """
    End-to-end process for answering queries about insurance policy PDFs using Bedrock.
    """
    chroma_client = PersistentClient(path=db_path)

    # Check if indexing is needed
    index_needed = force_reindex
    if not force_reindex:
        try:
            collection = chroma_client.get_collection(name=collection_name)
            if collection.count() == 0:
                print("Existing collection is empty, indexing is required.")
                index_needed = True
            else:
                 print(f"Using existing index with {collection.count()} chunks.")
                 index_needed = False
        except:
            print(f"Collection '{collection_name}' not found, indexing is required.")
            index_needed = True

    embedding_model = None

    # Extract and index if needed (same as before)
    if index_needed:
        print("Starting file indexing process...")
        if force_reindex:
            try:
                chroma_client.delete_collection(name=collection_name)
                print(f"Deleted existing collection '{collection_name}' for reindexing.")
            except:
                pass

        start_time = time.time()

        print("Extracting structured text from file...")
        structured_blocks = extract_structured_text_from_excel(pdf_path)
        print(f"  Extracted {len(structured_blocks)} text blocks.")
        if not structured_blocks:
             print("No text blocks extracted. Cannot proceed with indexing.")
             return {"query": query, "answer": "Error: Could not extract text from the file.", "context_chunks": [], "context_used": ""}

        print("Applying structural chunking...")
        structural_chunks = create_structural_chunks(structured_blocks)
        print(f"  Created {len(structural_chunks)} initial structural chunks.")

        print(f"Adding {CHUNK_OVERLAP} character overlap to chunks...")
        overlapping_chunks = add_overlap_to_chunks(structural_chunks, overlap=CHUNK_OVERLAP)
        print(f"  Resulting chunks after overlap processing: {len(overlapping_chunks)}")

        print("Storing chunks in vector database...")
        embedding_model = store_chunks_in_vector_db(
            overlapping_chunks,
            db_path=db_path,
            collection_name=collection_name,
            model_name=EMBEDDING_MODEL_NAME
        )
        index_time = time.time() - start_time
        print(f"Indexing completed in {index_time:.2f} seconds.")
    else:
        try:
             print(f"Loading embedding model: {EMBEDDING_MODEL_NAME}")
             embedding_model = SentenceTransformer(EMBEDDING_MODEL_NAME)
        except Exception as e:
             print(f"Error loading embedding model: {e}. Cannot process queries.")
             return {"query": query, "answer": "Error: Could not load the embedding model.", "context_chunks": [], "context_used": ""}

    # Process query (same retrieval logic)
    print(f"\nProcessing query: \"{query}\"")

    relevant_chunks = retrieve_relevant_chunks(
        query,
        embedding_model,
        db_path=db_path,
        collection_name=collection_name,
        rerank=True
    )

    if not relevant_chunks:
        print("No relevant chunks found for the query.")
        context = assemble_context(query, relevant_chunks)
        answer = generate_response_with_bedrock(query, context, bedrock_client)  # Use Bedrock
        return {
            "query": query,
            "answer": answer,
            "context_chunks": relevant_chunks,
            "context_used": context
        }

    # Assemble context for LLM
    context_token_budget = 30000 - (len(query.split()) * 1.3) - 1000  # Claude context window
    context = assemble_context(query, relevant_chunks, token_budget=int(context_token_budget))

    # Generate response using Bedrock
    answer = generate_response_with_bedrock(query, context, bedrock_client)  # Use Bedrock

    return {
        "query": query,
        "answer": answer,
        "context_chunks": relevant_chunks,
        "context_used": context
    }

# Add this function to your code
def sanitize_metadata(metadata):
    """
    Sanitize metadata to ensure all values are of accepted types (str, int, float, bool)
    for ChromaDB. Converts None values to empty strings.
    """
    sanitized = {}
    for key, value in metadata.items():
        if value is None:
            sanitized[key] = "" # Convert None to empty string
        elif isinstance(value, (str, int, float, bool)):
            sanitized[key] = value # Keep valid types as is
        else:
            sanitized[key] = str(value) # Convert other types to string
    return sanitized

def add_overlap_to_chunks(chunks: List[Dict[str, Any]], overlap: int = CHUNK_OVERLAP) -> List[Dict[str, Any]]:
    """
    Adds overlapping text from the end of a chunk to the beginning of the next.
    Considers section boundaries to avoid overlapping unrelated content.
    """
    overlapping_chunks = []

    for i, chunk in enumerate(chunks):
        overlapping_chunks.append(chunk) # Add the original chunk

        # Add overlap to the next chunk if it exists and is in the same section/subsection
        if i + 1 < len(chunks):
            next_chunk = chunks[i+1]

            # Check if chunks are in the same logical section/subsection
            same_section = (chunk["metadata"].get("section") == next_chunk["metadata"].get("section") and
                            chunk["metadata"].get("subsection") == next_chunk["metadata"].get("subsection"))
            # Also avoid overlapping across heading chunks themselves
            if not same_section or next_chunk["metadata"].get("type", "").startswith("heading_level_"):
                continue # Do not overlap across sections or into headings

            # Get overlap text from the end of the current chunk
            overlap_text = chunk["text"][-overlap:].strip()

            if overlap_text:
                 # Prepend overlap text to the beginning of the next chunk's text
                 # Ensure there's a clear separator
                 next_chunk["text"] = overlap_text + "\n---\n" + next_chunk["text"]

    return overlapping_chunks

# Part 9: Main Function
def main():
    """
    Main function to handle the flow using AWS Bedrock
    """
    # File path (same as before)
    pdf_path = "/home/yash/Desktop/fast/RFQ - Swara Baby products - 14 May 25 (1) (1).xlsx"
    
    # No need for local model path anymore
    db_path = "./insurance_policy_db"
    collection_name = "insurance_policy"
    force_reindex = False

    # Initialize Bedrock client
    bedrock_client = initialize_bedrock_client()
    if not bedrock_client:
        print("Failed to initialize Bedrock client. Exiting.")
        return

    # Validate file path
    if not os.path.exists(pdf_path):
        print(f"Error: File not found at {pdf_path}")
        return

    print(f"Insurance Policy Chat System with AWS Bedrock")
    print(f"================================================")
    print(f"File: {pdf_path}")
    print(f"Bedrock Model: {BEDROCK_MODEL_ID}")
    print(f"AWS Region: {AWS_REGION}")
    print(f"Vector Database: {db_path}")
    print(f"Collection: {collection_name}")
    print(f"Force Reindex: {force_reindex}")
    print(f"Embedding Model: {EMBEDDING_MODEL_NAME}")
    print(f"Reranker Model: {RERANKER_MODEL_NAME}")
    print(f"Chunk Target Size: {CHUNK_SIZE_TARGET} chars")
    print(f"Chunk Overlap: {CHUNK_OVERLAP} chars")
    print(f"================================================")

    # Process file indexing (if needed)
    print("Checking/Performing file indexing...")
    try:
        file_ext = os.path.splitext(pdf_path)[1].lower()
        
        if file_ext in ['.xlsx', '.xls']:
            print("Processing Excel file...")
            structured_blocks = extract_structured_text_from_excel(pdf_path)
            print(f"  Extracted {len(structured_blocks)} blocks from Excel.")
            
            if not structured_blocks:
                print("No blocks extracted. Cannot proceed with indexing.")
                return
                
            print("Applying structural chunking...")
            structural_chunks = create_structural_chunks(structured_blocks)
            print(f"  Created {len(structural_chunks)} structural chunks.")
            
            print(f"Adding {CHUNK_OVERLAP} character overlap to chunks...")
            overlapping_chunks = add_overlap_to_chunks(structural_chunks, overlap=CHUNK_OVERLAP)
            print(f"  Resulting chunks after overlap processing: {len(overlapping_chunks)}")
            
            print("Storing chunks in vector database...")
            if force_reindex:
                try:
                    chroma_client = PersistentClient(path=db_path)
                    chroma_client.delete_collection(name=collection_name)
                    print(f"Deleted existing collection '{collection_name}' for reindexing.")
                except Exception as e:
                    print(f"Note: {e}")
                    
            embedding_model = store_chunks_in_vector_db(
                overlapping_chunks,
                db_path=db_path,
                collection_name=collection_name,
                model_name=EMBEDDING_MODEL_NAME
            )
            print("Indexing completed successfully.")
            
        print("Indexing check/completion finished.")
    except Exception as e:
        print(f"An error occurred during the initial indexing check/process: {e}")
        import traceback
        traceback.print_exc()
        print("Please check the file path and ensure dependencies are installed.")
        return

    # Interactive query loop
    print("\nInsurance Policy Chat System Ready!")
    print("Type 'exit' or 'quit' to end the session")

    while True:
        query = input("\nEnter your question about the insurance policy: ")

        if query.lower() in ['exit', 'quit']:
            print("Exiting chat system. Goodbye!")
            break

        if not query.strip():
            continue

        start_time = time.time()
        try:
            result = process_insurance_pdf_query(
                pdf_path=pdf_path,
                query=query,
                bedrock_client=bedrock_client,  # Pass Bedrock client instead of model path
                db_path=db_path,
                collection_name=collection_name,
                force_reindex=False
            )

            query_time = time.time() - start_time

            print("\n" + "="*50)
            print("ANSWER:")
            print("-"*50)
            print(result["answer"])
            print("="*50)
            print(f"Response generated in {query_time:.2f} seconds using AWS Bedrock")

            show_context = input("\nWould you like to see the context used? (y/n): ")
            if show_context.lower() == 'y':
                print("\nCONTEXT USED:")
                print("-"*50)
                print(result["context_used"])
                print("-"*50)

        except Exception as e:
            print(f"\nAn error occurred while processing your query: {str(e)}")
            import traceback
            traceback.print_exc()

if __name__ == "__main__":
    main()